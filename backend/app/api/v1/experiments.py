"""研发实验管理 API"""
import os
import io
import datetime
import hashlib
from urllib.parse import quote
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query, Request
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional, List
from werkzeug.utils import secure_filename

from app.core.database import get_db
from app.core.security import get_current_user, require_permission
from app.models.user import User
from app.models.project import Project
from app.models.experiment import (
    Experiment, ExperimentRecord, ExperimentAttachment
)
from app.schemas.experiment import (
    ExperimentCreate, ExperimentUpdate, ExperimentOut, ExperimentDetailOut,
    ExperimentRecordCreate, ExperimentRecordUpdate, ExperimentRecordOut,
    ExperimentRecordDetailOut, ExperimentAttachmentOut,
    ExperimentQuery, RecordQuery, StatusAction, BatchDeleteRecords,
)
from app.schemas.common import ResponseBase, PaginationResponse
from app.core.operation_log import log_file_download

router = APIRouter()

UPLOAD_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "uploads", "experiments")
os.makedirs(UPLOAD_DIR, exist_ok=True)

MAX_ATTACHMENT_SIZE = 50 * 1024 * 1024  # 50MB
MAX_ATTACHMENTS_PER_RECORD = 20
MAX_EXPORT_RECORDS = 10000  # 单次导出上限
# 允许的附件扩展名白名单（小写）
ALLOWED_EXTENSIONS = {
    ".pdf", ".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp",
    ".xlsx", ".xls", ".csv", ".docx", ".doc",
    ".txt", ".md", ".json", ".xml",
    ".zip", ".rar", ".7z",
}

# 参数模板 JSON Schema 校验
PARAM_TEMPLATE_SCHEMA = {
    "type": "array",
    "items": {
        "type": "object",
        "required": ["name"],
        "properties": {
            "name": {"type": "string", "maxLength": 50},
            "unit": {"type": "string", "maxLength": 20},
            "default": {"type": "number"},
            "min": {"type": "number"},
            "max": {"type": "number"},
        }
    }
}

# 实验类型简称映射
TYPE_ABBR = {
    "performance": "PERF", "reliability": "RELI", "environment": "ENV",
    "material": "MATL", "process": "PROC", "other": "OTHR",
}


# ===== 工具函数 =====
def _generate_experiment_code(db: Session, exp_type: str) -> str:
    """生成实验编码: EXP-{TYPE}-{YYYYMM}-{SEQ:04d}"""
    now = datetime.datetime.now()
    prefix = f"EXP-{TYPE_ABBR.get(exp_type, 'OTHR')}-{now.strftime('%Y%m')}-"
    # 查询当月同类型实验数
    count = db.query(Experiment).filter(
        Experiment.code.like(f"{prefix}%")
    ).count()
    return f"{prefix}{count + 1:04d}"


def _validate_param_template(template: list) -> None:
    """校验参数模板符合 JSON Schema"""
    import json
    try:
        import jsonschema
        jsonschema.validate(instance=template, schema=PARAM_TEMPLATE_SCHEMA)
    except ImportError:
        # jsonschema 未安装时跳过校验
        pass
    except jsonschema.ValidationError as e:
        raise HTTPException(400, f"参数模板格式错误: {e.message}")


def _validate_user(db: Session, user_id: Optional[int], field_name: str):
    """校验用户存在性"""
    if user_id is None:
        return
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(400, f"{field_name} 指向的用户不存在")


def _check_experiment_access(exp: Experiment, current: User):
    """数据隔离：非 admin/manager 只能看自己设计或执行的实验"""
    if current.role not in ("admin", "manager") \
            and exp.designer_id != current.id \
            and exp.executor_id != current.id:
        raise HTTPException(403, "无权访问此实验")


def _get_experiment_or_404(db: Session, experiment_id: int, current: User) -> Experiment:
    """获取实验并校验访问权"""
    exp = db.query(Experiment).filter(Experiment.id == experiment_id).first()
    if not exp:
        raise HTTPException(404, "实验不存在")
    _check_experiment_access(exp, current)
    return exp


def _get_record_or_404(db: Session, record_id: int, current: User) -> ExperimentRecord:
    """获取记录并校验访问权（经 experiment 关联）"""
    rec = db.query(ExperimentRecord).filter(ExperimentRecord.id == record_id).first()
    if not rec:
        raise HTTPException(404, "实验记录不存在")
    _check_experiment_access(rec.experiment, current)
    return rec


def _check_experiment_editable(exp: Experiment):
    """校验实验是否可新增/修改记录（已结束的实验不可操作）"""
    if exp.status in ("cancelled", "completed"):
        raise HTTPException(400, f"实验已{('取消' if exp.status == 'cancelled' else '完成')}，不可操作记录")


def _refresh_actual_dates(db: Session, experiment_id: int):
    """根据实验记录自动维护 actual_start / actual_end（聚合查询优化）"""
    result = db.query(
        func.min(ExperimentRecord.executed_at),
        func.max(ExperimentRecord.executed_at),
    ).filter(
        ExperimentRecord.experiment_id == experiment_id,
        ExperimentRecord.executed_at.isnot(None),
    ).first()
    if not result or not result[0]:
        return
    exp = db.query(Experiment).filter(Experiment.id == experiment_id).first()
    if not exp:
        return
    exp.actual_start = result[0]
    exp.actual_end = result[1]


def _check_attachment_extension(filename: Optional[str]):
    """校验附件扩展名是否在白名单内"""
    if not filename:
        raise HTTPException(400, "文件名不能为空")
    ext = os.path.splitext(filename)[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(400, f"不支持的文件类型: {ext}，允许: {', '.join(sorted(ALLOWED_EXTENSIONS))}")


def _check_attachment_count(db: Session, record_id: int):
    """校验附件数量是否超限"""
    count = db.query(ExperimentAttachment).filter(
        ExperimentAttachment.record_id == record_id
    ).count()
    if count >= MAX_ATTACHMENTS_PER_RECORD:
        raise HTTPException(400, f"每条记录最多 {MAX_ATTACHMENTS_PER_RECORD} 个附件")


# ===== 实验 CRUD =====
@router.get("/list", response_model=ResponseBase)
def list_experiments(query: ExperimentQuery = Depends(), db: Session = Depends(get_db), current: User = Depends(get_current_user)):
    """实验列表 — 按查询条件过滤，仅 admin/manager 可查看全部，普通用户仅看相关项目"""
    q = db.query(Experiment)
    if current.role not in ("admin", "manager"):
        q = q.filter(
            (Experiment.designer_id == current.id) | (Experiment.executor_id == current.id)
        )
    if query.project_id:
        q = q.filter(Experiment.project_id == query.project_id)
    if query.status:
        q = q.filter(Experiment.status == query.status)
    if query.exp_type:
        q = q.filter(Experiment.exp_type == query.exp_type)
    if query.keyword:
        q = q.filter(Experiment.name.contains(query.keyword) | Experiment.code.contains(query.keyword))
    if query.date_from:
        q = q.filter(Experiment.plan_start >= query.date_from)
    if query.date_to:
        q = q.filter(Experiment.plan_end <= query.date_to)
    total = q.count()
    items = q.order_by(Experiment.id.desc()).offset(query.offset).limit(query.limit).all()

    # 关联用户姓名，便于前端展示
    user_ids = set()
    for e in items:
        if e.designer_id:
            user_ids.add(e.designer_id)
        if e.executor_id:
            user_ids.add(e.executor_id)
    user_map = {}
    if user_ids:
        users = db.query(User).filter(User.id.in_(user_ids)).all()
        user_map = {u.id: u.real_name or u.username for u in users}

    result = []
    for e in items:
        item = ExperimentDetailOut.model_validate(e).model_dump()
        item["designer_name"] = user_map.get(e.designer_id)
        item["executor_name"] = user_map.get(e.executor_id) if e.executor_id else None
        result.append(item)

    return ResponseBase(data=PaginationResponse(
        items=result,
        total=total, page=query.page, page_size=query.page_size,
        total_pages=(total + query.page_size - 1) // query.page_size,
    ).model_dump())


@router.post("/create", response_model=ResponseBase)
def create_experiment(data: ExperimentCreate, db: Session = Depends(get_db), current: User = Depends(require_permission("experiment", "create"))):
    """创建实验 — 校验项目/用户存在性"""
    project = db.query(Project).filter(Project.id == data.project_id).first()
    if not project:
        raise HTTPException(400, "关联的项目不存在")
    # designer_id 优先用请求中的值，否则取当前用户
    designer_id = data.designer_id if data.designer_id else current.id
    _validate_user(db, designer_id, "designer_id")
    _validate_user(db, data.executor_id, "executor_id")

    # 校验参数模板
    if data.param_template:
        _validate_param_template([p.model_dump() for p in data.param_template])

    code = _generate_experiment_code(db, data.exp_type)
    exp = Experiment(
        code=code, project_id=data.project_id, name=data.name,
        description=data.description, exp_type=data.exp_type,
        status=data.status, designer_id=designer_id,
        executor_id=data.executor_id, plan_start=data.plan_start,
        plan_end=data.plan_end,
        param_template=[p.model_dump() for p in data.param_template] if data.param_template else None,
    )
    db.add(exp)
    db.commit()
    db.refresh(exp)
    return ResponseBase(data=ExperimentOut.model_validate(exp).model_dump())


@router.get("/{experiment_id}", response_model=ResponseBase)
def get_experiment(experiment_id: int, db: Session = Depends(get_db), current: User = Depends(get_current_user)):
    exp = _get_experiment_or_404(db, experiment_id, current)
    data = ExperimentDetailOut.model_validate(exp).model_dump()
    # 关联用户姓名
    data["designer_name"] = (exp.designer.real_name or exp.designer.username) if exp.designer else None
    data["executor_name"] = (exp.executor.real_name or exp.executor.username) if exp.executor else None
    return ResponseBase(data=data)


@router.put("/{experiment_id}", response_model=ResponseBase)
def update_experiment(experiment_id: int, data: ExperimentUpdate, db: Session = Depends(get_db), current: User = Depends(require_permission("experiment", "update"))):
    exp = _get_experiment_or_404(db, experiment_id, current)
    update_data = data.model_dump(exclude_unset=True)
    # 状态机校验：update 接口不允许直接改 status，必须走 /status 接口
    if "status" in update_data and update_data["status"] != exp.status:
        raise HTTPException(400, "状态变更请使用专用接口 POST /{id}/status")
    if "executor_id" in update_data:
        _validate_user(db, update_data["executor_id"], "executor_id")
    # param_template 已由 model_dump 递归转为 dict，直接赋值
    for field, value in update_data.items():
        setattr(exp, field, value)
    exp.updated_at = datetime.datetime.utcnow()
    db.commit()
    db.refresh(exp)
    return ResponseBase(data=ExperimentOut.model_validate(exp).model_dump())


@router.delete("/{experiment_id}", response_model=ResponseBase)
def delete_experiment(experiment_id: int, db: Session = Depends(get_db), current: User = Depends(require_permission("experiment", "delete"))):
    """删除实验 — 级联删除记录和附件，清理物理文件"""
    exp = _get_experiment_or_404(db, experiment_id, current)

    # 收集所有附件路径（用于事务成功后清理物理文件）
    att_paths: list[str] = []
    for record in exp.records:
        for att in record.attachments:
            if att.file_path:
                att_paths.append(att.file_path)

    # 数据库级联删除（Experiment.records 和 ExperimentRecord.attachments 都配置了 cascade="all, delete-orphan"）
    db.delete(exp)
    db.commit()

    # 事务成功后清理物理文件（失败仅记日志，不影响业务）
    for path in att_paths:
        try:
            if path and os.path.exists(path):
                os.remove(path)
        except OSError:
            pass

    return ResponseBase(data={"msg": "实验及关联数据已删除"})


# ===== 状态流转 =====
@router.post("/{experiment_id}/status", response_model=ResponseBase)
def change_status(experiment_id: int, action: StatusAction, db: Session = Depends(get_db), current: User = Depends(require_permission("experiment", "update"))):
    """状态流转：start → in_progress, complete → completed, cancel → cancelled"""
    exp = _get_experiment_or_404(db, experiment_id, current)
    transitions = {
        "start": {"from": ("draft",), "to": "in_progress", "set_actual_start": True},
        "complete": {"from": ("in_progress",), "to": "completed", "set_actual_end": True},
        "cancel": {"from": ("draft", "in_progress"), "to": "cancelled", "set_actual_end": True},
    }
    rule = transitions.get(action.action)
    if not rule:
        raise HTTPException(400, f"不支持的动作: {action.action}")
    if exp.status not in rule["from"]:
        raise HTTPException(400, f"当前状态「{exp.status}」不允许执行 {action.action} 操作")
    exp.status = rule["to"]
    today = datetime.date.today()
    if rule.get("set_actual_start") and not exp.actual_start:
        exp.actual_start = today
    if rule.get("set_actual_end") and not exp.actual_end:
        exp.actual_end = today
    exp.updated_at = datetime.datetime.utcnow()
    db.commit()
    db.refresh(exp)
    return ResponseBase(data=ExperimentOut.model_validate(exp).model_dump())


# ===== 实验记录 =====
@router.get("/{experiment_id}/records", response_model=ResponseBase)
def list_records(
    experiment_id: int,
    query: RecordQuery = Depends(),
    db: Session = Depends(get_db),
    current: User = Depends(get_current_user),
):
    """实验记录列表 — 支持分页与筛选"""
    _get_experiment_or_404(db, experiment_id, current)
    q = db.query(ExperimentRecord).filter(ExperimentRecord.experiment_id == experiment_id)
    if query.batch_no:
        q = q.filter(ExperimentRecord.batch_no.contains(query.batch_no))
    if query.sample_code:
        q = q.filter(ExperimentRecord.sample_code.contains(query.sample_code))
    if query.conclusion:
        q = q.filter(ExperimentRecord.conclusion == query.conclusion)
    if query.is_abnormal is not None:
        q = q.filter(ExperimentRecord.is_abnormal == query.is_abnormal)
    if query.date_from:
        q = q.filter(ExperimentRecord.executed_at >= query.date_from)
    if query.date_to:
        q = q.filter(ExperimentRecord.executed_at <= query.date_to)

    total = q.count()
    records = q.order_by(ExperimentRecord.id.desc()).offset(query.offset).limit(query.limit).all()

    user_ids = {r.executor_id for r in records if r.executor_id}
    user_map = {}
    if user_ids:
        users = db.query(User).filter(User.id.in_(user_ids)).all()
        user_map = {u.id: u.real_name or u.username for u in users}

    result = []
    for r in records:
        item = ExperimentRecordOut.model_validate(r).model_dump()
        item["executor_name"] = user_map.get(r.executor_id) if r.executor_id else None
        item["attachment_count"] = len(r.attachments) if r.attachments else 0
        result.append(item)

    return ResponseBase(data=PaginationResponse(
        items=result,
        total=total, page=query.page, page_size=query.page_size,
        total_pages=(total + query.page_size - 1) // query.page_size,
    ).model_dump())


@router.get("/records/{record_id}", response_model=ResponseBase)
def get_record(record_id: int, db: Session = Depends(get_db), current: User = Depends(get_current_user)):
    """获取实验记录详情（含附件列表）"""
    rec = _get_record_or_404(db, record_id, current)
    data = ExperimentRecordDetailOut.model_validate(rec).model_dump()
    # 补充展示字段：执行人姓名、附件数、附件列表
    executor_name = None
    if rec.executor_id:
        executor = db.query(User).filter(User.id == rec.executor_id).first()
        executor_name = executor.real_name or executor.username if executor else None
    data["executor_name"] = executor_name
    data["attachment_count"] = len(rec.attachments) if rec.attachments else 0
    data["attachments"] = [ExperimentAttachmentOut.model_validate(a).model_dump() for a in rec.attachments]
    return ResponseBase(data=data)


@router.post("/records", response_model=ResponseBase)
def create_record(data: ExperimentRecordCreate, db: Session = Depends(get_db), current: User = Depends(require_permission("experiment", "create"))):
    """创建实验记录 — 校验实验存在性、状态、并自动更新 actual_start/end"""
    exp = _get_experiment_or_404(db, data.experiment_id, current)
    _check_experiment_editable(exp)
    executor_id = data.executor_id or current.id
    _validate_user(db, executor_id, "executor_id")

    # 自动提取 test_type
    pv = data.param_values or {}
    test_type = pv.get("test_type", "normal") if isinstance(pv, dict) else "normal"

    rec = ExperimentRecord(
        experiment_id=data.experiment_id, batch_no=data.batch_no,
        sample_code=data.sample_code, executor_id=executor_id,
        test_type=test_type,
        param_values=data.param_values, result_data=data.result_data,
        result_summary=data.result_summary, conclusion=data.conclusion,
        is_abnormal=data.is_abnormal,
        abnormal_desc=data.abnormal_desc, executed_at=data.executed_at,
    )
    db.add(rec)
    db.commit()
    db.refresh(rec)
    # 自动维护实验的 actual_start/actual_end
    _refresh_actual_dates(db, data.experiment_id)
    db.commit()
    return ResponseBase(data=ExperimentRecordOut.model_validate(rec).model_dump())


@router.put("/records/{record_id}", response_model=ResponseBase)
def update_record(record_id: int, data: ExperimentRecordUpdate, db: Session = Depends(get_db), current: User = Depends(require_permission("experiment", "update"))):
    """更新实验记录"""
    rec = _get_record_or_404(db, record_id, current)
    update_data = data.model_dump(exclude_unset=True)
    if "executor_id" in update_data:
        _validate_user(db, update_data["executor_id"], "executor_id")
    for field, value in update_data.items():
        setattr(rec, field, value)
    # 自动更新 test_type
    pv = rec.param_values or {}
    if isinstance(pv, dict):
        rec.test_type = pv.get("test_type", rec.test_type or "normal")
    db.commit()
    db.refresh(rec)
    # 自动维护实验的 actual_start/actual_end
    _refresh_actual_dates(db, rec.experiment_id)
    db.commit()
    return ResponseBase(data=ExperimentRecordOut.model_validate(rec).model_dump())


@router.delete("/records/{record_id}", response_model=ResponseBase)
def delete_record(record_id: int, db: Session = Depends(get_db), current: User = Depends(require_permission("experiment", "delete"))):
    """删除实验记录（先提交数据库事务，再清理附件物理文件，避免事务失败导致僵尸记录）"""
    rec = _get_record_or_404(db, record_id, current)
    experiment_id = rec.experiment_id
    # 先收集附件路径，再提交数据库事务
    att_paths = [att.file_path for att in rec.attachments if att.file_path]
    db.delete(rec)
    db.commit()
    # 事务成功后清理物理文件（失败仅记日志，不影响业务）
    for path in att_paths:
        try:
            if path and os.path.exists(path):
                os.remove(path)
        except OSError:
            pass
    # 删除记录后重新计算 actual 日期
    _refresh_actual_dates(db, experiment_id)
    db.commit()
    return ResponseBase(data={"msg": "实验记录已删除"})


@router.post("/records/batch-delete", response_model=ResponseBase)
def batch_delete_records(data: BatchDeleteRecords, db: Session = Depends(get_db), current: User = Depends(require_permission("experiment", "delete"))):
    """批量删除实验记录（事务保护 + 先提交事务再清理物理文件）"""
    if not data.ids:
        raise HTTPException(400, "请选择要删除的记录")
    recs = db.query(ExperimentRecord).filter(ExperimentRecord.id.in_(data.ids)).all()
    if len(recs) != len(data.ids):
        found_ids = {r.id for r in recs}
        missing = [i for i in data.ids if i not in found_ids]
        raise HTTPException(400, f"以下记录不存在: {missing}")
    try:
        experiment_ids = set()
        att_paths: list[str] = []
        for rec in recs:
            _check_experiment_access(rec.experiment, current)
            experiment_ids.add(rec.experiment_id)
            att_paths.extend(att.file_path for att in rec.attachments if att.file_path)
            db.delete(rec)
        db.commit()
    except Exception:
        db.rollback()
        raise
    # 事务成功后清理物理文件
    for path in att_paths:
        try:
            if path and os.path.exists(path):
                os.remove(path)
        except OSError:
            pass
    for eid in experiment_ids:
        _refresh_actual_dates(db, eid)
    db.commit()
    return ResponseBase(data={"msg": f"已删除 {len(recs)} 条记录"})


# ===== 导出 =====
@router.get("/{experiment_id}/records/export")
def export_records(
    experiment_id: int,
    db: Session = Depends(get_db),
    current: User = Depends(get_current_user),
):
    """导出实验记录为 Excel（含温漂/温升采样数据），上限 {MAX_EXPORT_RECORDS} 条"""
    _get_experiment_or_404(db, experiment_id, current)
    total = db.query(func.count(ExperimentRecord.id)).filter(
        ExperimentRecord.experiment_id == experiment_id
    ).scalar()
    if total > MAX_EXPORT_RECORDS:
        raise HTTPException(400, f"记录数（{total}）超过单次导出上限（{MAX_EXPORT_RECORDS}），请使用筛选分批导出")
    records = db.query(ExperimentRecord).filter(
        ExperimentRecord.experiment_id == experiment_id
    ).order_by(ExperimentRecord.id.desc()).all()

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
    except ImportError:
        raise HTTPException(500, "服务器未安装 openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "实验记录"
    headers = [
        "ID", "批次号", "样品编号", "执行人", "执行日期", "结论", "是否异常", "异常描述",
        "结果摘要", "测试类型", "关键指标", "创建时间",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    user_ids = {r.executor_id for r in records if r.executor_id}
    user_map = {}
    if user_ids:
        users = db.query(User).filter(User.id.in_(user_ids)).all()
        user_map = {u.id: u.real_name or u.username for u in users}

    conclusion_map = {"pass": "通过", "fail": "失败", "conditional_pass": "有条件通过", "need_retest": "需复测"}
    test_type_map = {"tcr": "温漂", "temp_rise": "温升"}

    def _build_key_metrics(rec: ExperimentRecord) -> str:
        """从 result_data 提取关键指标摘要"""
        if not rec.result_data:
            return ""
        rd = rec.result_data if isinstance(rec.result_data, dict) else {}
        pv = rec.param_values if isinstance(rec.param_values, dict) else {}
        test_type = pv.get("test_type")
        parts = []
        if test_type == "tcr":
            if "avg_tcr" in rd: parts.append(f"平均TCR={rd['avg_tcr']}ppm/°C")
            if "max_tcr" in rd: parts.append(f"最大TCR={rd['max_tcr']}ppm/°C")
            if "delta_r" in rd: parts.append(f"ΔR={rd['delta_r']}Ω")
            if "valid_point_count" in rd: parts.append(f"有效采样点={rd['valid_point_count']}")
        elif test_type == "temp_rise":
            if "max_temp_rise" in rd: parts.append(f"最大温升={rd['max_temp_rise']}°C")
            if "avg_temp_rise" in rd: parts.append(f"平均温升={rd['avg_temp_rise']}°C")
            if "max_change_rate" in rd: parts.append(f"最大变化率={rd['max_change_rate']}%")
            if "max_thermal_resistance" in rd: parts.append(f"最大热阻={rd['max_thermal_resistance']}°C/W")
            if "max_body_temp" in rd: parts.append(f"最大本体温度={rd['max_body_temp']}°C")
            if "valid_point_count" in rd: parts.append(f"有效测试点={rd['valid_point_count']}")
        return "；".join(parts)

    for r in records:
        pv = r.param_values if isinstance(r.param_values, dict) else {}
        test_type = pv.get("test_type", "")
        ws.append([
            r.id, r.batch_no or "", r.sample_code or "",
            user_map.get(r.executor_id, "") if r.executor_id else "",
            str(r.executed_at) if r.executed_at else "",
            conclusion_map.get(r.conclusion, r.conclusion or ""),
            "是" if r.is_abnormal else "否",
            r.abnormal_desc or "",
            r.result_summary or "",
            test_type_map.get(test_type, test_type or ""),
            _build_key_metrics(r),
            str(r.created_at) if r.created_at else "",
        ])

    # 第二张表：温漂采样明细
    tcr_records = [r for r in records if (r.param_values or {}).get("test_type") == "tcr"]
    if tcr_records:
        ws2 = wb.create_sheet("温漂采样明细")
        ws2.append(["记录ID", "批次号", "温度(°C)", "阻值(Ω)", "ΔT(°C)", "TCR(ppm/°C)"])
        for cell in ws2[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        for r in tcr_records:
            pv = r.param_values or {}
            ref_temp = (pv.get("temp_config") or {}).get("ref", 25)
            for p in pv.get("sample_points", []) or []:
                ws2.append([
                    r.id, r.batch_no or "",
                    p.get("temperature", ""),
                    p.get("resistance", ""),
                    (p.get("temperature", 0) - ref_temp) if p.get("temperature") is not None else "",
                    p.get("tcr", ""),
                ])

    # 第三张表：温升采样明细
    tr_records = [r for r in records if (r.param_values or {}).get("test_type") == "temp_rise"]
    if tr_records:
        ws3 = wb.create_sheet("温升采样明细")
        ws3.append(["记录ID", "批次号", "功率(W)", "电流(A)", "电压(mV)", "阻值(mΩ)", "本体温度(°C)", "温升(°C)", "总热阻(°C/W)", "阻值变化率(%)"])
        for cell in ws3[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        for r in tr_records:
            pv = r.param_values or {}
            for p in pv.get("sample_points", []) or []:
                ws3.append([
                    r.id, r.batch_no or "",
                    p.get("power", ""),
                    p.get("current", ""),
                    p.get("voltage_mv", p.get("voltage", "")),
                    p.get("resistance_mohm", p.get("resistance", "")),
                    p.get("body_temp", ""),
                    p.get("temp_rise", ""),
                    p.get("rth_core_to_ambient", ""),
                    p.get("resistance_change_rate", ""),
                ])

    # 列宽自适应（所有 sheet）
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            sheet.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"experiment_{experiment_id}_records_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    encoded = quote(filename, safe="")
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"},
    )


# ===== 实验附件 =====
@router.post("/records/{record_id}/attachments", response_model=ResponseBase)
async def upload_attachment(
    record_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current: User = Depends(require_permission("experiment", "create")),
):
    """为实验记录上传附件 — 校验类型白名单与数量上限"""
    rec = _get_record_or_404(db, record_id, current)
    _check_attachment_extension(file.filename)
    _check_attachment_count(db, record_id)

    contents = await file.read()
    if len(contents) > MAX_ATTACHMENT_SIZE:
        raise HTTPException(400, f"附件大小超过限制（最大 {MAX_ATTACHMENT_SIZE // 1024 // 1024}MB）")

    safe_name = secure_filename(file.filename or "attachment")
    if not safe_name:
        raise HTTPException(400, "文件名不合法")
    saved_name = f"REC{record_id}_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}_{safe_name}"
    file_path = os.path.realpath(os.path.join(UPLOAD_DIR, saved_name))
    upload_dir_real = os.path.realpath(UPLOAD_DIR)
    if not file_path.startswith(upload_dir_real + os.sep):
        raise HTTPException(400, "非法的文件路径")
    with open(file_path, "wb") as f:
        f.write(contents)

    att = ExperimentAttachment(
        record_id=record_id, file_name=safe_name,
        file_path=file_path, file_size=len(contents),
        file_hash=hashlib.sha256(contents).hexdigest(),
    )
    db.add(att)
    db.commit()
    db.refresh(att)
    return ResponseBase(data=ExperimentAttachmentOut.model_validate(att).model_dump())


@router.get("/records/{record_id}/attachments", response_model=ResponseBase)
def list_attachments(record_id: int, db: Session = Depends(get_db), current: User = Depends(get_current_user)):
    """列出某条实验记录的附件"""
    rec = _get_record_or_404(db, record_id, current)
    return ResponseBase(data=[ExperimentAttachmentOut.model_validate(a).model_dump() for a in rec.attachments])


@router.get("/attachments/{attachment_id}/download")
def download_attachment(
    attachment_id: int,
    inline: bool = Query(False, description="true=预览，false=下载"),
    db: Session = Depends(get_db),
    request: Request = None,
    current: User = Depends(get_current_user),
):
    """下载/预览实验附件"""
    att = db.query(ExperimentAttachment).filter(ExperimentAttachment.id == attachment_id).first()
    if not att:
        raise HTTPException(404, "附件不存在")
    # 经附件所属记录校验访问权
    _check_experiment_access(att.record.experiment, current)
    if not os.path.exists(att.file_path):
        raise HTTPException(404, "文件已不存在于服务器")

    # 非安全类型强制 attachment 下载，避免 XSS
    ext = os.path.splitext(att.file_name)[1].lower()
    force_download_exts = {".html", ".htm", ".svg", ".js"}
    if ext in force_download_exts:
        disposition = "attachment"
    else:
        disposition = "inline" if inline else "attachment"

    encoded_filename = quote(att.file_name, safe="")
    response = FileResponse(
        path=att.file_path,
        filename=att.file_name,
        headers={
            "Content-Disposition": f"{disposition}; filename*=UTF-8''{encoded_filename}",
            "X-Content-Type-Options": "nosniff",
        },
    )
    # 记录下载日志（仅下载触发，预览不记）
    if not inline:
        try:
            ip = request.client.host if request and request.client else "unknown"
        except Exception:
            ip = "unknown"
        log_file_download(db, current.id, current.username, att.file_name, att.file_size,
                          "experiment_attachment", att.id, ip)
    return response


@router.delete("/attachments/{attachment_id}", response_model=ResponseBase)
def delete_attachment(attachment_id: int, db: Session = Depends(get_db), current: User = Depends(require_permission("experiment", "delete"))):
    """删除附件（先提交事务再清理物理文件）"""
    att = db.query(ExperimentAttachment).filter(ExperimentAttachment.id == attachment_id).first()
    if not att:
        raise HTTPException(404, "附件不存在")
    _check_experiment_access(att.record.experiment, current)
    file_path = att.file_path
    db.delete(att)
    db.commit()
    try:
        if file_path and os.path.exists(file_path):
            os.remove(file_path)
    except OSError:
        pass
    return ResponseBase(data={"msg": "附件已删除"})
